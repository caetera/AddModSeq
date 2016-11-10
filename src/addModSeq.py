# -*- coding: utf-8 -*-
"""
Created on Sat Sep 13 13:06:18 2014

Modified sequence 

@author: vgor
"""
import sys
import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from os import path
from ttk import Frame, Label, Entry, Button 
from Tkinter import *
import tkFileDialog
import tkMessageBox

#constants
moddict = {}
modifications = {}
prsProb = 0.0

#regexes
daExpr = re.compile("N[^P][STC]")
flankExpr = re.compile(r"\[(.+)\]\.(\w+)\.\[(.+)\]")
modExpr = re.compile(r"([\w-]+)\((\w+)\)")
ptmRSExpr = re.compile(r"(\w+)\((\D+)\):\s*([\d.,]+)")
phRSExpr = re.compile(r"(\w+)\((\d+)\)(?:x\d+)?:\s*([\d.,]+)")

#GUI
class ListBoxChoice(object):
    def __init__(self, master=None, title=None, message=None, data = []):
        self.master = master
        self.value = None
        self.data = data
        
        self.modalPane = Toplevel(self.master)

        self.modalPane.transient(self.master)
        self.modalPane.grab_set()

        self.modalPane.bind("<Return>", self._choose)
        self.modalPane.bind("<Escape>", self._cancel)
        
        self.modalPane.columnconfigure(0, pad = 5, weight = 1)
        
        self.modalPane.rowconfigure(0, pad = 5)
        self.modalPane.rowconfigure(1, pad = 5, weight = 1)
        self.modalPane.rowconfigure(2, pad = 5)

        if title:
            self.modalPane.title(title)

        if message:
            Label(self.modalPane, text=message).grid(row = 0, column = 0, padx = 10, sticky = W+E)

        listFrame = Frame(self.modalPane)
        listFrame.grid(row = 1, column = 0, sticky = W+E)
        
        scrollBar = Scrollbar(listFrame)
        scrollBar.pack(side=RIGHT, fill=Y)
        self.listBox = Listbox(listFrame, selectmode = SINGLE)
        self.listBox.pack(side = LEFT, fill = BOTH, expand = True)
        scrollBar.config(command = self.listBox.yview)
        self.listBox.config(yscrollcommand=scrollBar.set)
        for item in self.data:
            self.listBox.insert(END, item)

        buttonFrame = Frame(self.modalPane)
        buttonFrame.grid(row = 2, column = 0, sticky = W+E)

        chooseButton = Button(buttonFrame, text="Choose", command=self._choose)
        chooseButton.pack(padx=5, pady=5, side = LEFT)

        cancelButton = Button(buttonFrame, text="Cancel", command=self._cancel)
        cancelButton.pack(padx=5, pady=5, side=RIGHT)

    def _choose(self, event=None):
        try:
            firstIndex = self.listBox.curselection()[0]
            self.value = self.data[int(firstIndex)]
        except IndexError:
            self.value = None
        self.modalPane.destroy()

    def _cancel(self, event=None):
        self.modalPane.destroy()
        
    def returnValue(self):
        self.master.wait_window(self.modalPane)
        return self.value

class toolUI(Frame):
  
    def __init__(self, parent):
        Frame.__init__(self, parent)   
         
        self.parent = parent
        
        self.initUI()        
        
    def initUI(self):
      
        self.parent.title("Sequence modification parser")          
        
        self.columnconfigure(0, pad=5)
        self.columnconfigure(1, pad=5, weight = 1)
        self.columnconfigure(2, pad=5)
        
        self.rowconfigure(0, pad=5, weight = 1)
        self.rowconfigure(1, pad=5, weight = 1)
        self.rowconfigure(2, pad=5, weight = 1)
        self.rowconfigure(3, pad=5, weight = 1)
        self.rowconfigure(4, pad=5, weight = 1)
        self.rowconfigure(5, pad=5, weight = 1)
        
        self.lInput = Label(self, text = "Input file")
        self.lInput.grid(row = 0, column = 0, sticky = W)
        self.lPRS = Label(self, text = "phospoRS Column Name")
        self.lPRS.grid(row = 1, column = 0, sticky = W)
        self.lScore = Label(self, text = "Min phosphoRS Score")
        self.lScore.grid(row = 2, column = 0, sticky = W)
        self.lDA = Label(self, text = "Check deamidation sites")
        self.lDA.grid(row = 3, column = 0, sticky = W)
        self.lLabFile = Label(self, text = "Label description file")
        self.lLabFile.grid(row = 4, column = 0, sticky = W)
        
        self.ifPathText = StringVar()
        self.prsScoreText = StringVar(value = '0')
        self.prsNameText = StringVar()
        self.doDAVar = IntVar()
        self.labFileText = StringVar(value = path.abspath('moddict.txt'))
        
        self.ifPath = Entry(self, textvariable = self.ifPathText)
        self.ifPath.grid(row = 0, column = 1, sticky = W+E)
        self.prsName = Entry(self, textvariable = self.prsNameText)
        self.prsName.grid(row = 1, column = 1, sticky = W+E)        
        self.prsScore = Entry(self, textvariable = self.prsScoreText, state = DISABLED)
        self.prsScore.grid(row = 2, column = 1, sticky = W+E)
        self.doDABox = Checkbutton(self, variable = self.doDAVar)
        self.doDABox.grid(row = 3, column = 1, sticky = W)
        self.labFile = Entry(self, textvariable = self.labFileText)
        self.labFile.grid(row = 4, column = 1, sticky = W+E)
        
        
        self.foButton = Button(self, text = "Select file", command = self.selectFileOpen)
        self.foButton.grid(row = 0, column = 2, sticky = E+W, padx = 3)
        self.prsButton = Button(self, text = "Select column", state = DISABLED, command = self.selectColumn)
        self.prsButton.grid(row = 1, column = 2, sticky = E+W, padx = 3)
        self.labFileButton = Button(self, text = "Select file", command = self.selectLabFileOpen)
        self.labFileButton.grid(row = 4, column = 2, sticky = E+W, padx = 3)
        self.startButton = Button(self, text = "Start", command = self.start, padx = 3, pady = 3)
        self.startButton.grid(row = 5, column = 1, sticky = E+W)
        
    
        self.pack(fill = BOTH, expand = True, padx = 5, pady = 5)
        
    def selectFileOpen(self):
        #Open file dialog
        dlg = tkFileDialog.Open(self, filetypes = [('Excel spreadsheet', '*.xlsx')])
        openName = dlg.show()
        if openName != "":
            self.ifPathText.set(openName)
            self.findPRS(openName)#find phosphRS column
    
    def selectLabFileOpen(self):
        #Open labels file dialog
        dlg = tkFileDialog.Open(self, filetypes = [('All files', '*.*')])
        openName = dlg.show()
        if openName != "":
            self.labFileText.set(openName)
    
    def findPRS(self, openName):
        #find phosphoRS column
        worksheet = load_workbook(openName, use_iterators = True).get_active_sheet() #open excel sheet
        
        self.headers = [unicode(worksheet.cell(get_column_letter(columnNr) + '1').value).lower()
            for columnNr in range(1, worksheet.get_highest_column() + 1)]
        
        self.prsButton.config(state = "normal")
        
        if 'phosphors site probabilities' in self.headers:
            self.prsNameText.set('phosphoRS Site Probabilities')
        elif 'phosphors: phospho_sty site probabilities' in self.headers:
            self.prsNameText.set('PhosphoRS: Phospho_STY Site Probabilities')
        elif 'phosphors: phospho site probabilities' in self.headers:
            self.prsNameText.set('PhosphoRS: Phospho Site Probabilities')
        else:
            self.prsNameText.set('PhosphoRS column not found')
            return
            
        self.prsScore.config(state = 'normal')
        if self.prsScoreText.get() == '0':
            self.prsScoreText.set('95')
    
    def selectColumn(self):
        column = ListBoxChoice(self.parent, "Select column",
                                    "Select the column that has phosphoRS or ptmRS data", self.headers).returnValue()
        
        if not column is None:
            self.prsNameText.set(column)
            self.prsScore.config(state = 'normal')
            if self.prsScoreText.get() == '0':
                self.prsScoreText.set('95')
    
    def start(self):
        if self.prsScoreText.get() == '':
            self.prsScoreText.set("0")
            
        process(self.ifPathText.get(), float(self.prsScoreText.get()), bool(self.doDAVar.get()),
                'PD', self.labFileText.get(), self.prsNameText.get())

class ModificationGroup(object):
    """
    Represent one specific type of modifications, ex. Phosphorylation or Oxidation etc
    """
    def __init__(self, name, symbol):
        self.name = name
        self.symbol = symbol
        self.unbound = 0
        self.positions = []
    
    def __repr__(self):
        return "{}<{}>: BoundTo: {} + {} unbound".\
            format(self.name, self.symbol, " ".join(self.positions), self.unbound)

class Peptide(object):
    """
    Represent peptide
    """
    def  __init__(self, sequence, flankN = "", flankC = ""):
        self.sequence = sequence
        self.modifications = {}
        self.mutations = []
        self.flankN = flankN
        self.flankC = flankC
        
    def setModification(self, name, modification):
        self.modifications[name] = modification
    
    def addMutation(self, mutation):
        self.mutations.append(mutation)
    
    def iterateFlanks(self, useN = True, useC = True):
        #return all possible sequences with different flanking residues
        #useN - use N-terminal flanks
        #useC - use C-terminal flanks
        #each letter in flanking string represent possible flanking AA
        if useN and len(self.flankN) > 0:
            nAA = [a for a in self.flankN]
        else:
            nAA = [""]
            
        if useC and len(self.flankC) > 0:
            cAA = [a for a in self.flankC]
        else:
            cAA = [""]
            
        #here we should work with mutated sequence
        return (n + self._applyMutations() + c for n in nAA for c in cAA)
    
    def verifyDaSites(self):
        #trying to deduce the name for deamidaton
        daName = ""
        for testName in modifications["Deamidation"]:
            if self.modifications.has_key(testName):
                daName = testName
                break
                
        if daName == "":#no demidation
            return 
 
        #collect valid deamidation sites
        validSites = set()
        for testSequence in self.iterateFlanks(useN = False):
            for site in daExpr.finditer(testSequence):
                validSites.add(site.start() + 1)
                    
        #check if ptmRS assigned sites properly    
        for loc in self.modifications[daName].positions[:]:
            pos = int(loc[1:])
            if not pos in validSites:
                #remove invalid
                self.modifications[daName].positions.remove(loc)
                self.modifications[daName].unbound += 1
            else:
                #remove properly assigned sites from the set of available sites
                validSites.remove(pos)
        
        #if the number of unbound modifications is greater or equal to number 
        #of available sites assign them all
        if len(validSites) <= self.modifications[daName].unbound:
            locs = ["{}{}".format(self.sequence[s - 1], s) for s in validSites]
            self.modifications[daName].positions.extend(locs)
            self.modifications[daName].unbound -= len(validSites)
        
    def toModX(self):
        result = ""
        modText = defaultdict(list)
        mutSequence = self._applyMutations()
        
        for name in sorted(self.modifications.keys()):
            for loc in self.modifications[name].positions:
                if loc in ("N-Term", "C-Term"):
                    modText[loc].append(self.modifications[name].symbol)
                else:
                    code = loc[0]
                    pos = int(loc[1:]) - 1
                    self._checkLocation(code, pos)
                    modText[pos].append(self.modifications[name].symbol)
            
            modText["Unbound"].append("({})".format(self.modifications[name].symbol)\
                                                    * self.modifications[name].unbound)
        
        if len(modText["Unbound"]) > 0:
            result += "{}".format("".join(modText["Unbound"]))
            
        if len(modText["N-term"]) > 0:
            result += "{}-".format(",".join(modText["N-term"]))
            
        for i in range(len(mutSequence)):
            result += "{}{}".format(",".join(modText[i]), mutSequence[i])
        
        if len(modText["C-term"]) > 0:
            result += "-{}".format(",".join(modText["C-term"]))        
        
        return "[{}].{}.[{}]".format(self.flankN, result, self.flankC)
    
    def _checkLocation(self, code, pos):
        if self.sequence[pos] != code:
            raise Exception("Modification position mismatch in sequence '{}' at #{}:\
                            {} is not equal to {}".format(self.sequence, pos + 1,\
                            self.sequence[pos], code))
    
    def _applyMutations(self):
        result = self.sequence
        for loc, mutAA in self.mutations:
            code = loc[0]
            pos = int(loc[1:]) - 1
            
            if result[pos] == code:
                result = result[:pos] + mutAA + result[pos+1:]
            else:
                raise Exception("Mutation mismatch in sequence '{}' at #{}:\
                                {} is not equal to {}".format(result, pos + 1,\
                                result[pos], code))
        
        return result
    
    def __repr__(self):
        return "Peptide: [{}].{}.[{}]\nModifications:\n{}\nMutations: {}"\
                .format(self.flankN, self.sequence, self.flankC,\
                "\n".join(map(repr, self.modifications.values())),\
                self.mutations)

def parseModDict(moddictInput):
    """
    Reads text file and updates dictionary of modifications
    Parameters
    moddictInput - path or filelike object to read dictionary from
    Return
    None
    """
    with open(moddictInput) as fin:
        for line in fin:
            if line.startswith('#') or line.strip() == '': #comments
                continue
            else: 
                parts = line.strip().split('\t')
                try:
                    if parts[1] == 'NONE':
                        parts[1] = ''
                    modifications[parts[0]] = map(lambda s: s.strip(), parts[2].split(','))#all possible representaton of a modification
                    for r in modifications[parts[0]]:
                        moddict[r] = parts[1]
                except:
                    raise Exception('Error parsing modifications on the following line:\n{}'.format(line))

def analyzePRS(prsString):
    """
    Convert ptmRS string into modification dictionary
    """
    result = defaultdict(list)
    for item in ptmRSExpr.finditer(prsString):
        pos, name, prob = item.groups()
        try:
            prob = float(prob)
        except ValueError:
            prob = float(prob.replace(",", "."))
            
        result[name].append((pos, prob))
    
    return result

def analyzePhRS(phrsString):
    """
    Convert phosphoRS string into modification dictionary
    """
    result = defaultdict(list)
    for item in phRSExpr.finditer(phrsString):
        aa, pos, prob = item.groups()
        try:
            prob = float(prob)
        except ValueError:
            prob = float(prob.replace(",", "."))
            
        result["Phosp"].append((aa+pos, prob)) #5 letter code is to be compatible with ptmRS
    
    return result


def analyzeMod(modString):
    """
    Convert modification string into modification dictionary
    and the list of amino acid mutations
    """
    mods = defaultdict(list)
    mutations = []
    for item in modExpr.finditer(modString):
        pos, name = item.groups()
        
        #split modifications from mutations
        if pos[0] in ("X", "B", "Z", "J") and len(name) == 1 and name.isupper():
            mutations.append((pos, name))
        else:
            mods[name].append(pos)
    
    return (mods, mutations)

def createPeptide(seqString, modString, prsString, minProb):
    """
    Convert a combination of sequence string, modification string and ptmRS string
    into single Peptide object.
    """
    try: #match to sequence with flanking residues
        flankN, sequence, flankC = flankExpr.match(seqString).groups()
        result = Peptide(sequence.upper(), flankN, flankC)
    except AttributeError: #if match was unsuccessful this exception is fired
        result = Peptide(seqString.upper())
    
    modDict, mutations = analyzeMod(modString)
    if prsString != "":       
        prsDict = analyzePRS(prsString) #try parsing with ptmRS rules
        if prsDict == {}: #if parsing was unsuccessful, i.e. no modifications found
            prsDict = analyzePhRS(prsString) #try with phosphoRS rules
        if prsDict == {}:
            print "WARNING Can not parse modification string: {}".format(prsString)
    else:
        prsDict = {}
    
    result.mutations.extend(mutations)
     
    for name in prsDict.iterkeys(): #first check PRS assignments
        modItem = filter(lambda (k, v): k[:5] == name, modDict.iteritems())
        #Since ptmRS uses first 5 letters of modification name as identifier
        #we have to find the full name in the modString
        #Assumingly there should be only one match, but...
        if len(modItem) > 1:
            raise Exception("Ambiguous modification name in ptmRS string:\n\
                             '{}' vs {}".format(name, [i[0] for i in modItem]))
        elif len(modItem) == 0:
            raise Exception("Unidentified modification in PRS string: {}".format(name))
        
        modItem = modItem[0] #remove corresponding element from modDict
        
        boundPositions = [p for p, pp in filter(lambda (p, pp): pp >= minProb, prsDict[name])]
        modification = ModificationGroup(modItem[0], moddict[modItem[0]])
        modification.positions.extend(boundPositions)
        modification.unbound = len(modItem[1]) - len(boundPositions)

        result.setModification(modification.name, modification)
        
        modDict.pop(modItem[0])
    
    for name in modDict.iterkeys(): #add all modifications, that were not assigned by ptmRS
        modification = ModificationGroup(name, moddict[name])
        modification.positions.extend(modDict[name])
        
        result.setModification(modification.name, modification)
    
    return result

#functions
def parseCLInput(arguments):
    """
    Parse input from command line
    """
    raise NotImplementedError("Yet to be updated")
    
    if not path.isfile(arguments[1]):
            print "Can't find filename: {}".format(arguments[1])
            return False
            
    if len(arguments) == 2:
        print "Using default minimal phosphoRS site probability (95)\nUsing usual deamidation treatment\nUsing default inputMode (PD)\nUsing default labels (moddict.txt)"
        arguments = arguments[1:2] + ['95.0', 'N', 'PD', 'moddict.txt']
    elif len(arguments) == 3:
        print "Using usual deamidation treatment\nUsing default inputMode (PD)\nUsing default labels (moddict.txt)"
        arguments = arguments[1:3] + ['N', 'PD', 'moddict.txt']
    elif len(arguments) == 4:
        print "Using default inputMode (PD)\nUsing default labels (moddict.txt)"
        arguments = arguments[1:4] + ['PD', 'moddict.txt']
    elif len(arguments) == 5:
        print "Using default labels (moddict.txt)"
        arguments = arguments[1:5] + ['moddict.txt']
    elif len(arguments) >= 6:
        if not path.isfile(arguments[5]):
                print "Can't find filename: {}".format(arguments[5])
                return False
        arguments = arguments[1:6]
    
    arguments[1] = float(arguments[1])
    if arguments[2].upper() == 'Y':
        arguments[2] = True
    elif arguments[2].upper() == 'N':
        arguments[2] = False
    else:
        print "Can't parse deamidation treatment: {}".format(arguments[2])
        return False
    
    return arguments

def printUsage():
    """
    Print usage info
    """
    raise NotImplementedError("Yet to be updated")
    
    print 'Usage: {} excelInputFile [minPhosphoRS] [doDeamidation] [inputMode] [labels]\n\
    \tminPhosphoRS (optional) minimal phospoRS site probability to consider valid phospo site\n\
    \tdoDeamidation (optional) perform check for deamidation sites, can be Y(es) or N(o)\n\
    \tinputMode (optional) can be PD (Proteome Discoverer) or MQ (MaxQuant)\n\
    \tlabels (optional) should be path to file with label descriptions'.format(sys.argv[0])

def hitEnter():
    """
    Promt to hit enter
    """
    raw_input('\nHit ENTER to continue')

def runInteractive():
    """
    Interactive mode
    """
    raise NotImplementedError("Yet to be updated")
    
    print "Hi, I'm here to help you with modifications.\nUsually I'm happy to run with some command line arguments, here is the small summary"
    printUsage()
    print "Unfortunately you have not provided any arguments. Don't worry, I'll guide you through"
    
    inputFile = raw_input('Please, type the path of the excel file you want to process, ex. somefile.xlsx\nThis is the only manadatory parameter\n')
    while not path.isfile(inputFile):
        inputFile = raw_input('Sorry, I can not find this file. Please, try again\n')
    
    minPRS = raw_input('Please, type the minimal phosphoRS site probability to consider valid phospho site(0 - 100)\n\
    If the input file does not contain phospoRS information the value will be ignored\nThe default settings is 95, you happy with it just hit ENTER\n')
    if minPRS == '':
        minPRS = 95.0
    else:
        minPRS = float(minPRS)
        if minPRS < 0 or minPRS > 100:
            print 'WARNING! Minimal phospoRS site probability is expected to be from 0 to 100'
    
    doDA = raw_input('Please, select if you require validation of deamidation sites\n\tY for Yes and N for No\nThe dafault is No, if you happy with it just hit ENTER\n')
    while not doDA.upper() in ['Y', 'N', '']:
        doDA = raw_input('Sorry, I can not understand. Please, try again\n')
    if doDA == '':
        doDA = 'N'
    
    inputMode = raw_input('Please, type the file type\n\tPD for Proteome Discoverer and MQ for MaxQuant\nThe dafault is PD, if you happy with it just hit ENTER\n')
    while not inputMode.upper() in ['MQ', 'PD', '']:
        inputMode = raw_input('Sorry, I can not understand. Please, try again\n')
    if inputMode == '':
        inputMode = 'PD'
            
    modFile = raw_input('Please, type the path of the modifications file\nThe default is moddict.txt if you happy with this just hit ENTER\n')
    while not (path.isfile(modFile) or modFile == ''):
        modFile = raw_input('Sorry, I can not find this file. Please, try again\n')
    if modFile == '':
        if not path.isfile('moddict.txt'):
            raise Exception('Default file does not exist')
        modFile = 'moddict.txt'
    
    
    print 'Nice. Thank you.\nIf you want to run the same analysis in future you can use the following command in console:\n'
    print '{} {} {} {} {} {}'.format(sys.argv[0], inputFile, minPRS, doDA, inputMode, modFile)
    hitEnter()
    
    if doDA.upper() == 'Y':
        doDA = True
    else:
        doDA = False
    
    process(inputFile, minPRS, doDA, inputMode, modFile)

def runGUI():
    """
    Run with GUI
    """
    try:
        root = Tk()
        ex = toolUI(root)
        root.geometry("500x250+200+200")
        root.mainloop()
    
    except Exception as ex:
        tkMessageBox.showerror('Error', ex.message)
        
def writeRow(worksheet, ColNr, rowNr, iterable):
    """
    Write row of values to the worksheet
    Parameters
    worksheet - openpyxl worksheet
    Colnr - starting column to write into
    rowNr - the number of row to write into
    iterable - any iterable of values
    Return
    None
    """
    for index in range(len(iterable)):
        worksheet.cell(get_column_letter(ColNr + index) + str(rowNr)).value = iterable[index]
            
def applyModsMQ(sequence):
    """
    Parse modified sequence in MaxQuant format to modX format
    Global dictionary moddict contains conversion rules
    """
    raise NotImplementedError("Yet to be updated")
    
    openbrace = sequence.find('(')
    closebrace = sequence.find(')')
    
    if openbrace == -1 and closebrace == -1:#no more mods
        return sequence[1:-1]#remove first and last underscores on return
        
    elif openbrace != -1 and closebrace != -1:#complete mod
        mod = sequence[openbrace + 1 : closebrace]
        res = sequence[:openbrace -1] + moddict[mod] + sequence[openbrace - 1] + sequence[closebrace + 1:]
        return applyModsMQ(res)
    
    else: #incomplete mod
        raise Exception('Invalid input sequence: {}'.format(sequence))

def process(excelInput, minPRS, doDA, inputMode, moddictInput, prsColumnName = None):
    """
    Reads Excel sheet and add modifiedSequence column in the end
    Parameters:
    excelInput - search results exported to Excel file, as provided by PD or MQ
        string with path or file-like object
    minPRS - minimal phospoRS site probability
        float
    doDA - check deamidation sites using conservative motiff
        bool
    inputMode - input format of excel file PD for Proteome Discoverer or MQ for MaxQuant
        string
    moddictInput - path to the textfile with modification labels
        string
    prsColumnName - name of the column with ptmRS results
        string or None
    Return:
    None
    """
    doPRS = False
    selectedColumns = {}
    
    parseModDict(moddictInput)
    
    print 'Opening excel file...'
        
    workbook = load_workbook(excelInput)#open worksheet to save result
    worksheet = workbook.get_active_sheet()
    lastColNr = worksheet.get_highest_column() + 1 #number of the first free column in the worksheet
    lastRowNr = worksheet.get_highest_row()
    
    for columnNr in range(1, worksheet.get_highest_column() + 1): #read column headers
        selectedColumns[unicode(worksheet.cell(get_column_letter(columnNr) + '1').value).lower()] = columnNr
    
    headers = ['Modified Sequence']
    
    if prsColumnName is None:
        prsColumnName = u'phosphors site probabilities'
    else:
        prsColumnName = prsColumnName.lower()
    
    if selectedColumns.has_key(prsColumnName): #check for phospoRS input
        print '\tphospoRS results was found\n\tphosopho sites will be assigned correspondingly'
        doPRS = True
        headers[0] += ' (PRS)'
    else:
        print '\tphospoRS results was NOT found\n\tphosopho sites will be assigned by modification column'
        
    if doDA:
        print '\tdeamidation sites will be checked'
        headers[0] += ' (DA)'
        
        
    writeRow(worksheet, lastColNr, 1, headers) #write headers
        
    writeCount = 0 #count number of lines written
    
    print 'Start processing excel file...'
    
    for rowNr in range(2, lastRowNr + 1):#read all except headers
        try:
            if doPRS:#PRS present
                PRSstring = worksheet.cell(get_column_letter(selectedColumns[prsColumnName]) + str(rowNr)).value
                if PRSstring is None:#empty cells
                    PRSstring = ""
            else:
                PRSstring = ""
            
            if worksheet.cell(get_column_letter(selectedColumns[u'sequence']) + str(rowNr)).value != None:
                seqString = worksheet.cell(get_column_letter(selectedColumns[u'sequence']) + str(rowNr)).value
                modString = worksheet.cell(get_column_letter(selectedColumns[u'modifications']) + str(rowNr)).value
                
                peptide = createPeptide(seqString, modString, PRSstring, minPRS)
                
                if doDA:
                    peptide.verifyDaSites()
                
                sequence = peptide.toModX()
                
            else:
                sequence = None
                
        except KeyError as ex:
            raise Exception("Missing column in the input file: {}".format(ex.message))
        
        writeRow(worksheet, lastColNr, rowNr, [sequence])
        writeCount += 1
        
        if writeCount % 1000 == 0:
            print '\r{} of {} lines ready.'.format(writeCount, lastRowNr - 1),
        
    print '\r\n{} of {} lines ready.\nSaving excel file...'.format(writeCount, lastRowNr - 1) 
        
    workbook.save(excelInput)
    
    print 'Done. It was pleasure to work for you.'

if __name__ == "__main__":
    if len(sys.argv) > 1:
        print 'Sequence modification parser'
        parameters = parseCLInput(sys.argv)
        if parameters:
            process(*parameters)
        else:
            printUsage()
            hitEnter()

    else:
        runGUI()
        #runInteractive()
